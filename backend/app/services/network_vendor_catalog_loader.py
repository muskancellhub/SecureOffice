from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from app.core.exceptions import AppError

NETWORK_VENDOR_CATALOG_FILENAME = 'network_vendor_catalog_public_pricing_meraki_extreme_skymirr_inhand.xlsx'
NETWORK_VENDOR_CATALOG_SOURCE_NAME = 'network_vendor_catalog_public_pricing_meraki_extreme_skymirr_inhand'
NETWORK_VENDOR_SHEET_NAME = 'Catalog_Public_Pricing'
NETWORK_VENDOR_ALLOWED_VENDORS = {'Meraki', 'Extreme Networks', 'SkyMirr', 'InHand'}

EXCLUDED_DEVICE_TERMS = ('laptop', 'phone', 'smartphone', 'tablet', 'mobile')

CATEGORY_MAP = {
    'wireless ap': 'wifi_ap',
    'switch': 'switch',
    'security / sd-wan': 'security_appliance',
    'cellular gateway': 'cellular_gateway',
    'cellular router': 'router',
    'fixed wireless access': 'router',
    'industrial router': 'router',
    'enterprise router': 'router',
    'iot edge gateway': 'cellular_gateway',
    'industrial switch': 'switch',
    'smart camera': 'camera',
    'sensor': 'sensor',
    'antenna': 'antenna',
    '5g router/cpe': 'router',
}


@dataclass
class NetworkVendorCatalogLoadResult:
    rows: list[dict[str, Any]]
    errors: list[str]
    skipped_count: int


def default_network_vendor_catalog_path() -> Path:
    # /backend/app/services -> /SecureOffice2 -> db/<file>
    return Path(__file__).resolve().parents[3] / 'db' / NETWORK_VENDOR_CATALOG_FILENAME


def _slugify(value: str) -> str:
    cleaned = re.sub(r'[^a-zA-Z0-9]+', '-', (value or '').strip().lower())
    return cleaned.strip('-') or 'unknown'


def _to_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_price(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).strip()
    if not raw:
        return None

    lowered = raw.lower()
    if lowered in {'n/a', 'na', '-', '--', 'tbd', 'quote', 'quote only'}:
        return None

    normalized = raw.replace(',', '').replace('$', '').strip()
    try:
        return float(normalized)
    except ValueError:
        return None


def _is_excluded_device(category: str, family_type: str, model: str) -> bool:
    text = f'{category} {family_type} {model}'.lower()
    return any(term in text for term in EXCLUDED_DEVICE_TERMS)


def _normalize_category(raw_category: str, family_type: str) -> str:
    category = (raw_category or '').strip().lower()
    if category in CATEGORY_MAP:
        return CATEGORY_MAP[category]

    family = (family_type or '').lower()
    if 'switch' in family:
        return 'switch'
    if 'camera' in family:
        return 'camera'
    if 'sensor' in family:
        return 'sensor'
    if 'antenna' in family:
        return 'antenna'
    if 'firewall' in family or 'sd-wan' in family or 'security' in family:
        return 'security_appliance'
    if 'gateway' in family:
        return 'cellular_gateway'
    if 'router' in family or 'cpe' in family:
        return 'router'
    if 'access point' in family or 'wi-fi' in family or 'wifi' in family:
        return 'wifi_ap'
    return 'accessory'


def normalize_network_vendor_row(row: dict[str, Any], *, row_number: int) -> dict[str, Any] | None:
    vendor = _to_text(row.get('Vendor'))
    category_raw = _to_text(row.get('Category'))
    model = _to_text(row.get('Model'))
    family_type = _to_text(row.get('Family/Type'))
    currency = (_to_text(row.get('Currency')) or 'USD').upper()
    pricing_basis = _to_text(row.get('Pricing basis'))
    official_source = _to_text(row.get('Official catalog source'))
    public_price_source = _to_text(row.get('Public price source'))
    notes = _to_text(row.get('Notes'))

    if not vendor and not category_raw and not model:
        return None

    if vendor not in NETWORK_VENDOR_ALLOWED_VENDORS:
        return None

    if _is_excluded_device(category_raw, family_type, model):
        return None

    normalized_category = _normalize_category(category_raw, family_type)

    sku_model = model or f'row-{row_number}'
    sku = f"EXCEL-{_slugify(vendor)}-{_slugify(sku_model)}"
    name_parts = [vendor, model or family_type or normalized_category.replace('_', ' ').title()]
    name = ' '.join(part for part in name_parts if part).strip()

    price = _parse_price(row.get('Price'))

    attributes = {
        'category': normalized_category,
        'product_type': normalized_category,
        'brand': vendor,
        'model': model,
        'family_type': family_type,
        'pricing_basis': pricing_basis,
        'official_catalog_source': official_source,
        'public_price_source': public_price_source,
        'notes': notes,
        'source_type': 'excel',
        'source_name': NETWORK_VENDOR_CATALOG_SOURCE_NAME,
        'raw_category': category_raw,
        'raw_row_number': row_number,
    }

    return {
        'sku': sku,
        'name': name,
        'vendor': vendor,
        'vendor_sku': model or sku,
        'description': family_type or category_raw or None,
        'price': price,
        'currency': currency,
        'availability': 'in_stock',
        'attributes': attributes,
    }


def load_network_vendor_catalog(file_path: str | Path | None = None) -> NetworkVendorCatalogLoadResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - covered by import runtime behavior
        raise AppError('openpyxl is required for Excel catalog ingestion', 500) from exc

    workbook_path = Path(file_path) if file_path else default_network_vendor_catalog_path()
    if not workbook_path.exists():
        raise AppError(f'Network vendor catalog file not found: {workbook_path}', 404)

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if NETWORK_VENDOR_SHEET_NAME not in wb.sheetnames:
        raise AppError(
            f"Workbook missing expected sheet '{NETWORK_VENDOR_SHEET_NAME}'",
            400,
        )

    ws = wb[NETWORK_VENDOR_SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise AppError('Network vendor catalog sheet is empty', 400)

    headers = [(_to_text(cell) or f'column_{idx}') for idx, cell in enumerate(header)]

    normalized_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    skipped_count = 0

    for row_index, raw_row in enumerate(rows_iter, start=2):
        row = {headers[idx]: raw_row[idx] for idx in range(min(len(headers), len(raw_row)))}
        try:
            normalized = normalize_network_vendor_row(row, row_number=row_index)
            if not normalized:
                skipped_count += 1
                continue
            normalized_rows.append(normalized)
        except Exception as exc:
            errors.append(f'row {row_index}: {exc}')

    wb.close()
    return NetworkVendorCatalogLoadResult(rows=normalized_rows, errors=errors, skipped_count=skipped_count)
